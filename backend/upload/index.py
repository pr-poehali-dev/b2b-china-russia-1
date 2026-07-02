import json
import os
import base64
import uuid
import boto3
import psycopg2
from psycopg2.extras import RealDictCursor

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token',
    'Access-Control-Max-Age': '86400',
    'Content-Type': 'application/json',
}

SCHEMA = 't_p93404869_b2b_china_russia_1'


def _db():
    dsn = os.environ['DATABASE_URL']
    sep = '&' if '?' in dsn else '?'
    dsn += f'{sep}options=-csearch_path%3D{SCHEMA}'
    conn = psycopg2.connect(dsn)
    conn.autocommit = True
    return conn, conn.cursor(cursor_factory=RealDictCursor)


def _s3():
    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )


def _cdn(key: str) -> str:
    return f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"


def _resp(status: int, body: dict):
    return {'statusCode': status, 'headers': CORS, 'isBase64Encoded': False, 'body': json.dumps(body, default=str)}


def _parse_token(token: str):
    import hashlib, hmac
    SECRET = os.environ.get('DATABASE_URL', 'sinobridge-salt')
    try:
        raw = base64.urlsafe_b64decode(token.encode()).decode()
        parts = raw.split('.')
        seller_id, ts, sig = parts[0], parts[1], parts[2]
        payload = f"{seller_id}.{ts}"
        expected = hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:24]
        if hmac.compare_digest(sig, expected):
            return int(seller_id)
    except Exception:
        return None
    return None


def handler(event: dict, context) -> dict:
    '''Загрузка фото товаров в S3 и импорт товаров из Excel/CSV.'''
    method = event.get('httpMethod', 'POST')
    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'isBase64Encoded': False, 'body': ''}

    headers = event.get('headers') or {}
    token = headers.get('X-Auth-Token') or headers.get('x-auth-token', '')
    seller_id = _parse_token(token) if token else None
    if not seller_id:
        return _resp(401, {'error': 'Требуется авторизация'})

    params = event.get('queryStringParameters') or {}
    action = params.get('action', 'photo')

    body_raw = event.get('body') or ''
    if event.get('isBase64Encoded'):
        body_raw = base64.b64decode(body_raw).decode('utf-8', errors='replace')

    body = {}
    try:
        body = json.loads(body_raw)
    except Exception:
        return _resp(400, {'error': 'Неверный формат JSON'})

    s3 = _s3()

    # --- UPLOAD PHOTO ---
    if action == 'photo':
        file_b64 = body.get('file')
        content_type = body.get('content_type', 'image/jpeg')
        if not file_b64:
            return _resp(400, {'error': 'Нет файла'})

        try:
            file_bytes = base64.b64decode(file_b64)
        except Exception:
            return _resp(400, {'error': 'Неверный base64'})

        ext_map = {'image/jpeg': 'jpg', 'image/png': 'png', 'image/webp': 'webp', 'image/gif': 'gif'}
        ext = ext_map.get(content_type, 'jpg')
        key = f"products/{seller_id}/{uuid.uuid4().hex}.{ext}"

        s3.put_object(Bucket='files', Key=key, Body=file_bytes, ContentType=content_type)
        return _resp(200, {'url': _cdn(key)})

    # --- IMPORT FROM EXCEL/CSV ---
    if action == 'import_excel':
        import io, csv
        file_b64 = body.get('file')
        filename = body.get('filename', 'import.csv')
        if not file_b64:
            return _resp(400, {'error': 'Нет файла'})

        try:
            file_bytes = base64.b64decode(file_b64)
        except Exception:
            return _resp(400, {'error': 'Неверный base64'})

        ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else 'csv'
        rows = []

        if ext in ('xlsx', 'xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                headers_row = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers_row, [str(v or '').strip() for v in row])))
            except Exception as e:
                return _resp(400, {'error': f'Ошибка чтения Excel: {e}'})
        else:
            try:
                text = file_bytes.decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
            except Exception as e:
                return _resp(400, {'error': f'Ошибка чтения CSV: {e}'})

        if not rows:
            return _resp(400, {'error': 'Файл пустой или нет данных'})

        # Column name aliases
        def pick(row, *keys):
            for k in keys:
                if k in row and row[k]:
                    return row[k]
            return ''

        conn, cur = _db()
        imported = []
        errors = []
        try:
            for i, row in enumerate(rows[:200]):
                name = pick(row, 'название', 'name', 'наименование', 'товар', 'product')
                if not name:
                    errors.append(f'Строка {i+2}: нет названия')
                    continue
                category = pick(row, 'категория', 'category')
                price = pick(row, 'цена', 'price', 'стоимость')
                description = pick(row, 'описание', 'description')
                image_url = pick(row, 'фото', 'photo', 'image', 'image_url', 'url')
                sku = pick(row, 'артикул', 'sku', 'код')
                min_order = pick(row, 'мин. заказ', 'min_order', 'минимальный заказ')
                quantity = pick(row, 'количество', 'quantity', 'остаток')

                cur.execute(
                    "INSERT INTO products (seller_id, name, category, price, description, image_url, sku, min_order, quantity) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                    "RETURNING id, name, category, price, description, image_url, status, views, sku, min_order, quantity, created_at",
                    (seller_id, name, category or None, price or None, description or None,
                     image_url or None, sku or None, min_order or None, quantity or None)
                )
                imported.append(dict(cur.fetchone()))
        finally:
            cur.close()
            conn.close()

        return _resp(200, {
            'imported': imported,
            'count': len(imported),
            'errors': errors,
        })

    # --- UPLOAD VIDEO (прямой put_object, лимит ~7 МБ base64) ---
    if action == 'video_upload':
        file_b64 = body.get('file')
        content_type = body.get('content_type', 'video/mp4')
        if not file_b64:
            return _resp(400, {'error': 'Нет файла'})
        try:
            file_bytes = base64.b64decode(file_b64)
        except Exception:
            return _resp(400, {'error': 'Неверный base64'})
        max_bytes = 7 * 1024 * 1024
        if len(file_bytes) > max_bytes:
            return _resp(413, {'error': f'Файл слишком большой: {len(file_bytes)//1024//1024} МБ. Максимум 7 МБ.'})
        ext_map = {
            'video/mp4': 'mp4', 'video/quicktime': 'mov',
            'video/x-msvideo': 'avi', 'video/webm': 'webm', 'video/mpeg': 'mpg',
        }
        ext = ext_map.get(content_type, 'mp4')
        key = f"media/{seller_id}/videos/{uuid.uuid4().hex}.{ext}"
        s3.put_object(Bucket='files', Key=key, Body=file_bytes, ContentType=content_type)
        return _resp(200, {'url': _cdn(key)})

    return _resp(404, {'error': 'Неизвестное действие'})