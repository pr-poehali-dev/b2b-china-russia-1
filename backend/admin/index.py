import json
import os
import hashlib
import hmac
import base64
import time
import uuid
import psycopg2
import boto3
from psycopg2.extras import RealDictCursor

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Token',
    'Access-Control-Max-Age': '86400',
    'Content-Type': 'application/json',
}

SCHEMA = 't_p93404869_b2b_china_russia_1'
SECRET = os.environ.get('DATABASE_URL', 'sinobridge-admin-salt')
ADMIN_PASSWORD = os.environ.get('ADMIN_PANEL_PASSWORD', '')


def _db():
    dsn = os.environ['DATABASE_URL']
    sep = '&' if '?' in dsn else '?'
    dsn += f'{sep}options=-csearch_path%3D{SCHEMA}'
    conn = psycopg2.connect(dsn)
    conn.autocommit = True
    return conn, conn.cursor(cursor_factory=RealDictCursor)


def _s3():
    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )


def _cdn(key: str) -> str:
    return f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"


def _get_catalog_seller_id(cur) -> int:
    '''Возвращает id системного поставщика-каталога (создаёт при первом обращении).'''
    cur.execute("SELECT id FROM sellers WHERE is_catalog = TRUE LIMIT 1")
    row = cur.fetchone()
    if row:
        return row['id']
    pw_hash = hashlib.sha256((uuid.uuid4().hex + 'sb').encode()).hexdigest()
    cur.execute(
        "INSERT INTO sellers (company_name, email, password_hash, plan, is_catalog) "
        "VALUES ('Каталог сайта', %s, %s, 'Verified', TRUE) RETURNING id",
        (f"catalog-{uuid.uuid4().hex[:10]}@internal.local", pw_hash),
    )
    return cur.fetchone()['id']


def _resp(status: int, body: dict):
    return {'statusCode': status, 'headers': CORS, 'isBase64Encoded': False, 'body': json.dumps(body, default=str)}


def _make_token() -> str:
    payload = f"admin.{int(time.time())}"
    sig = hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:32]
    raw = f"{payload}.{sig}"
    return base64.urlsafe_b64encode(raw.encode()).decode()


def _check_token(token: str) -> bool:
    try:
        raw = base64.urlsafe_b64decode(token.encode()).decode()
        parts = raw.split('.')
        payload = f"{parts[0]}.{parts[1]}"
        sig = parts[2]
        expected = hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:32]
        return hmac.compare_digest(sig, expected)
    except Exception:
        return False


def handler(event: dict, context) -> dict:
    '''Админ-панель: вход по паролю, управление товарами и поставщиками (создание, удаление, список).'''
    method = event.get('httpMethod', 'GET')
    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'isBase64Encoded': False, 'body': ''}

    params = event.get('queryStringParameters') or {}
    action = params.get('action', '')

    body = {}
    if event.get('body'):
        try:
            body = json.loads(event['body'])
        except Exception:
            body = {}

    headers = event.get('headers') or {}
    token = headers.get('X-Admin-Token') or headers.get('x-admin-token', '')

    # --- LOGIN ---
    if action == 'login' and method == 'POST':
        password = body.get('password') or ''
        if not ADMIN_PASSWORD or password != ADMIN_PASSWORD:
            return _resp(401, {'error': 'Неверный пароль'})
        return _resp(200, {'token': _make_token()})

    if not _check_token(token):
        return _resp(401, {'error': 'Требуется авторизация администратора'})

    conn, cur = _db()
    try:
        # --- SELLERS LIST ---
        if action == 'sellers' and method == 'GET':
            cur.execute(
                "SELECT id, company_name, email, province, category, plan, rating, "
                "reviews_count, created_at FROM sellers WHERE is_catalog IS NOT TRUE ORDER BY created_at DESC"
            )
            return _resp(200, {'sellers': [dict(s) for s in cur.fetchall()]})

        # --- CREATE SELLER ---
        if action == 'add_seller' and method == 'POST':
            company = (body.get('company_name') or '').strip()
            email = (body.get('email') or '').strip().lower()
            if not company or not email:
                return _resp(400, {'error': 'Укажите название компании и email'})
            cur.execute("SELECT id FROM sellers WHERE email = %s", (email,))
            if cur.fetchone():
                return _resp(409, {'error': 'Email уже зарегистрирован'})
            password = body.get('password') or base64.urlsafe_b64encode(os.urandom(9)).decode()
            pw_hash = hashlib.sha256((password + 'sb').encode()).hexdigest()
            cur.execute(
                "INSERT INTO sellers (company_name, email, password_hash, province, category, plan) "
                "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, company_name, email, province, category, plan, created_at",
                (company, email, pw_hash, body.get('province'), body.get('category'), body.get('plan', 'Verified')),
            )
            seller = dict(cur.fetchone())
            seller['generated_password'] = password if not body.get('password') else None
            return _resp(200, {'seller': seller})

        # --- DELETE SELLER (+ все связанные данные) ---
        if action == 'delete_seller' and method == 'POST':
            sid = body.get('id')
            if not sid:
                return _resp(400, {'error': 'Укажите id'})
            cur.execute("SELECT id FROM buyer_chats WHERE seller_id = %s", (sid,))
            chat_ids = [r['id'] for r in cur.fetchall()]
            if chat_ids:
                cur.execute("DELETE FROM buyer_chat_messages WHERE chat_id = ANY(%s)", (chat_ids,))
                cur.execute("DELETE FROM buyer_chats WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM buyer_notifications WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM leads WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM certificates WHERE seller_id = %s", (sid,))
            cur.execute("SELECT id FROM media WHERE seller_id = %s", (sid,))
            media_ids = [r['id'] for r in cur.fetchall()]
            if media_ids:
                cur.execute("DELETE FROM video_likes WHERE media_id = ANY(%s)", (media_ids,))
                cur.execute("DELETE FROM video_views WHERE media_id = ANY(%s)", (media_ids,))
            cur.execute("DELETE FROM video_views WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM media WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM premium_orders WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM products WHERE seller_id = %s", (sid,))
            cur.execute("DELETE FROM sellers WHERE id = %s", (sid,))
            return _resp(200, {'ok': True})

        # --- PRODUCTS LIST ---
        if action == 'products' and method == 'GET':
            cur.execute(
                "SELECT p.id, p.name, p.category, p.price, p.image_url, p.status, p.views, p.created_at, "
                "p.seller_id, s.company_name FROM products p "
                "JOIN sellers s ON s.id = p.seller_id "
                "ORDER BY p.created_at DESC"
            )
            return _resp(200, {'products': [dict(p) for p in cur.fetchall()]})

        # --- ADD PRODUCT (в общий каталог сайта) ---
        if action == 'add_product' and method == 'POST':
            name = (body.get('name') or '').strip()
            if not name:
                return _resp(400, {'error': 'Укажите название товара'})
            seller_id = _get_catalog_seller_id(cur)
            photos = body.get('photos', [])
            image_url = photos[0] if photos else body.get('image_url')
            cur.execute(
                "INSERT INTO products (seller_id, name, category, price, description, image_url, photos, sku, min_order, quantity) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                "RETURNING id, name, category, price, image_url, status, views, created_at",
                (seller_id, name, body.get('category'), body.get('price'), body.get('description'),
                 image_url, photos, body.get('sku'), body.get('min_order'), body.get('quantity')),
            )
            product = dict(cur.fetchone())
            product['seller_id'] = seller_id
            product['company_name'] = 'Каталог сайта'
            return _resp(200, {'product': product})

        # --- DELETE PRODUCT (полностью) ---
        if action == 'delete_product' and method == 'POST':
            pid = body.get('id')
            if not pid:
                return _resp(400, {'error': 'Укажите id'})
            cur.execute("UPDATE leads SET product_id = NULL WHERE product_id = %s", (pid,))
            cur.execute("UPDATE media SET product_id = NULL WHERE product_id = %s", (pid,))
            cur.execute("DELETE FROM products WHERE id = %s", (pid,))
            return _resp(200, {'ok': True})

        # --- UPLOAD PHOTO (для товаров каталога) ---
        if action == 'upload_photo' and method == 'POST':
            file_b64 = body.get('file')
            content_type = body.get('content_type', 'image/jpeg')
            if not file_b64:
                return _resp(400, {'error': 'Нет файла'})
            try:
                file_bytes = base64.b64decode(file_b64)
            except Exception:
                return _resp(400, {'error': 'Неверный base64'})
            ext_map = {'image/jpeg': 'jpg', 'image/png': 'png', 'image/webp': 'webp', 'image/gif': 'gif'}
            ext = ext_map.get(content_type, 'jpg')
            key = f"products/catalog/{uuid.uuid4().hex}.{ext}"
            s3 = _s3()
            s3.put_object(Bucket='files', Key=key, Body=file_bytes, ContentType=content_type)
            return _resp(200, {'url': _cdn(key)})

        # --- IMPORT PRODUCTS FROM EXCEL/CSV (в общий каталог сайта) ---
        if action == 'import_excel' and method == 'POST':
            import io, csv
            file_b64 = body.get('file')
            filename = body.get('filename', 'import.csv')
            if not file_b64:
                return _resp(400, {'error': 'Нет файла'})
            try:
                file_bytes = base64.b64decode(file_b64)
            except Exception:
                return _resp(400, {'error': 'Неверный base64'})

            ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else 'csv'
            rows = []

            if ext in ('xlsx', 'xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    ws = wb.active
                    headers_row = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers_row, [str(v or '').strip() for v in row])))
                except Exception as e:
                    return _resp(400, {'error': f'Ошибка чтения Excel: {e}'})
            else:
                try:
                    text = file_bytes.decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(text))
                    for row in reader:
                        rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
                except Exception as e:
                    return _resp(400, {'error': f'Ошибка чтения CSV: {e}'})

            if not rows:
                return _resp(400, {'error': 'Файл пустой или нет данных'})

            def pick(row, *keys):
                for k in keys:
                    if k in row and row[k]:
                        return row[k]
                return ''

            seller_id = _get_catalog_seller_id(cur)
            imported = []
            errors = []
            for i, row in enumerate(rows[:200]):
                name = pick(row, 'название', 'name', 'наименование', 'товар', 'product')
                if not name:
                    errors.append(f'Строка {i+2}: нет названия')
                    continue
                category = pick(row, 'категория', 'category')
                price = pick(row, 'цена', 'price', 'стоимость')
                description = pick(row, 'описание', 'description')
                image_url = pick(row, 'фото', 'photo', 'image', 'image_url', 'url')
                sku = pick(row, 'артикул', 'sku', 'код')
                min_order = pick(row, 'мин. заказ', 'min_order', 'минимальный заказ')
                quantity = pick(row, 'количество', 'quantity', 'остаток')

                cur.execute(
                    "INSERT INTO products (seller_id, name, category, price, description, image_url, sku, min_order, quantity) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                    "RETURNING id, name, category, price, description, image_url, status, views, sku, min_order, quantity, created_at",
                    (seller_id, name, category or None, price or None, description or None,
                     image_url or None, sku or None, min_order or None, quantity or None)
                )
                p = dict(cur.fetchone())
                p['seller_id'] = seller_id
                p['company_name'] = 'Каталог сайта'
                imported.append(p)

            return _resp(200, {'imported': imported, 'count': len(imported), 'errors': errors})

        # --- ADD LOGISTICS (КАРГО) ---
        if action == 'add_logistics' and method == 'POST':
            company = (body.get('company_name') or '').strip()
            ltype = (body.get('type') or '').strip()
            if not company or not ltype:
                return _resp(400, {'error': 'Укажите название компании и тип'})
            cur.execute(
                "INSERT INTO logistics (company_name, logo_url, type, description, routes, transit_time, "
                "min_weight, phone, email, website, telegram, wechat, featured) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                "RETURNING id, company_name, logo_url, type, description, routes, transit_time, "
                "min_weight, phone, email, website, telegram, wechat, rating, reviews_count, featured, created_at",
                (company, body.get('logo_url'), ltype, body.get('description'), body.get('routes'),
                 body.get('transit_time'), body.get('min_weight'), body.get('phone'), body.get('email'),
                 body.get('website'), body.get('telegram'), body.get('wechat'), bool(body.get('featured', False))),
            )
            return _resp(200, {'logistics': dict(cur.fetchone())})

        # --- LOGISTICS (КАРГО) LIST ---
        if action == 'logistics' and method == 'GET':
            cur.execute(
                "SELECT id, company_name, logo_url, type, description, routes, transit_time, "
                "min_weight, phone, email, website, telegram, wechat, rating, reviews_count, featured, created_at "
                "FROM logistics ORDER BY created_at DESC"
            )
            return _resp(200, {'logistics': [dict(r) for r in cur.fetchall()]})

        # --- DELETE LOGISTICS (КАРГО) ---
        if action == 'delete_logistics' and method == 'POST':
            lid = body.get('id')
            if not lid:
                return _resp(400, {'error': 'Укажите id'})
            cur.execute("DELETE FROM logistics_reviews WHERE logistics_id = %s", (lid,))
            cur.execute("DELETE FROM logistics WHERE id = %s", (lid,))
            return _resp(200, {'ok': True})

        return _resp(404, {'error': 'Неизвестное действие'})
    finally:
        cur.close()
        conn.close()